from tkinter import *
from tkinter import messagebox
import openpyxl
from datetime import datetime
import os
from openpyxl import Workbook

window = Tk()

window.title("معانی کلمات و جملات ترکی")
EXCEL_FILE = "Dict-TU-FA.xlsx"
current_row = 0

# Track if we are editing an existing entry
edit_mode = False
edit_row_number = None

# Source options (initial list - users can add more)
source_options = [
    "کتاب",      # Book
    "یوتیوب",    # YouTube
]

# Variables to store selected source and description
selected_source = StringVar()
selected_source.set(None)  # Nothing selected at first
source_description = StringVar(value="")

# Turkish special letters (6 lowercase + 6 uppercase = 12 letters)
turkish_letters = [
    'ç', 'ğ', 'ı', 'ö', 'ş', 'ü',  # lowercase
    'Ç', 'Ğ', 'İ', 'Ö', 'Ş', 'Ü'   # uppercase
]

# Function to insert letter at cursor position in the currently focused text box
def insert_letter(letter):
    """Insert Turkish special letter into the currently focused text box"""
    
    # Find which widget currently has focus (cursor)
    focused_widget = window.focus_get()
    
    # List of all text boxes that should accept Turkish letters
    turkish_text_boxes = [text_box_turk, text_box_pron, text_box_exam, text_box_syn, text_box_ant]
    
    # Check if the focused widget is one of our text boxes
    if focused_widget in turkish_text_boxes:
        try:
            cursor_pos = focused_widget.index(INSERT)
            focused_widget.insert(cursor_pos, letter)
        except:
            focused_widget.insert(END, letter)
        focused_widget.focus_set()
    else:
        # If no valid text box has focus, default to Turkish word box
        try:
            cursor_pos = text_box_turk.index(INSERT)
            text_box_turk.insert(cursor_pos, letter)
        except:
            text_box_turk.insert(END, letter)
        text_box_turk.focus_set()

def add_new_source():
    # Create pop-up window
    popup = Toplevel(window)
    popup.title("افزودن منبع جدید")
    popup.geometry("400x180")
    popup.resizable(False, False)
    
    # Center the popup on the main window
    popup.transient(window)
    popup.grab_set()
    
    # Label and Entry for new source
    label = Label(popup, text="منبع جدید را وارد کنید:", font=("B Nazanin", 12))
    label.pack(pady=20)
    
    new_source_entry = Entry(popup, font=("B Nazanin", 11), width=30)
    new_source_entry.pack(pady=10)
    new_source_entry.focus()
    
    # Function to add the source and close popup
    def add_and_close():
        new_source = new_source_entry.get().strip()
        
        if not new_source:
            messagebox.showwarning("اخطار", "لطفاً نام منبع را وارد کنید", parent=popup)
            return
        
        # Check if source already exists
        if new_source in source_options:
            messagebox.showwarning("اخطار", f"منبع '{new_source}' قبلاً وجود دارد!", parent=popup)
            return
        
        # Add to source_options list
        source_options.append(new_source)
        
        # Clear existing radio buttons
        for widget in radio_frame.winfo_children():
            widget.destroy()
        
        # Recreate all radio buttons (old + new)
        radio_buttons.clear()
        for i, option in enumerate(source_options):
            radio_btn = Radiobutton(
                radio_frame,
                text=option,
                font=("B Nazanin", 11),
                variable=selected_source,
                value=option,
                indicatoron=1,
                selectcolor="white",
                relief="flat"
            )
            radio_btn.grid(row=0, column=i, padx=5)
            radio_buttons.append(radio_btn)
            
        # Close the popup
        popup.destroy()

    # Button frame for equal-sized buttons
    button_frame = Frame(popup)
    button_frame.pack(pady=15)
        
    # Add button
    add_btn = Button(
        button_frame,
        text="افزودن",
        font=("B Nazanin", 10, "bold"),
        bg="#2E7D32",
        fg="white",
        activebackground="#1B5E20",
        width=12,
        padx=5,
        pady=5,
        command=add_and_close
        )
    add_btn.grid(row=0, column=0, padx=10)
        
    # Cancel button
    cancel_btn = Button(
        button_frame,
        text="انصراف",
        font=("B Nazanin", 10, "bold"),
        bg="#f44336",
        fg="white",
        activebackground="#d32f2f",
        width=12,
        padx=5,
        pady=5,
        command=popup.destroy
        )
    cancel_btn.grid(row=0, column=1, padx=10)
        
    # Bind Enter key to add_and_close function
    new_source_entry.bind('<Return>', lambda event: add_and_close())
        
    # Bind Escape key to close popup
    popup.bind('<Escape>', lambda event: popup.destroy())

def clear_all():
    global edit_mode, edit_row_number
    edit_mode = False
    edit_row_number = None
    
    text_box_turk.delete("1.0", END)
    text_box_def.delete("1.0", END)
    text_box_pron.delete("1.0", END)
    text_box_exam.delete("1.0", END)
    text_box_syn.delete("1.0", END)
    text_box_ant.delete("1.0", END)
    source_entry.delete(0, END)
    source_description.set("")
    selected_source.set(None)
    text_box_turk.focus_set()

# Function to remove selected source
def remove_selected_source():
    selected = selected_source.get()
    
    # Don't allow removing default sources (کتاب and یوتیوب)
    if selected == "کتاب" or selected == "یوتیوب":
        messagebox.showwarning("اخطار", "نمی‌توان منبع پیش‌فرض را حذف کرد")
        return
    
    if not selected or selected == "None":
        messagebox.showwarning("اخطار", "لطفاً ابتدا یک منبع را انتخاب کنید")
        return
    
    # Remove from source_options list
    source_options.remove(selected)
    
    # Clear existing radio buttons
    for widget in radio_frame.winfo_children():
        widget.destroy()
    
    # Recreate all radio buttons (remaining sources)
    radio_buttons.clear()
    for i, option in enumerate(source_options):
        radio_btn = Radiobutton(
            radio_frame,
            text=option,
            font=("B Nazanin", 11),
            variable=selected_source,
            value=option,
            indicatoron=1,
            selectcolor="white",
            relief="flat"
        )
        radio_btn.grid(row=0, column=i, padx=5)
        radio_buttons.append(radio_btn)
    
    # Deselect the removed source
    selected_source.set(None)
    
    # Show confirmation
    messagebox.showinfo("موفق", f"منبع '{selected}' با موفقیت حذف شد")

# ============= EXCEL FUNCTIONS =============

def get_next_row_number():
    """Get the next row number (count of existing entries + 1)"""
    if not os.path.exists(EXCEL_FILE):
        return 1  # First entry is row 1
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Count entries in column 3 (Turkish word) - ignore empty formatted rows
    entry_count = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=3).value is not None:
            entry_count += 1
    
    wb.close()
    return entry_count + 1

def get_current_date():
    """Get current date in English format (YYYY-MM-DD)"""
    return datetime.now().strftime("%Y-%m-%d")

def clear_form():
    """Clear all input fields after saving"""
    global edit_mode, edit_row_number
    edit_mode = False
    edit_row_number = None
    
    text_box_turk.delete("1.0", END)
    text_box_def.delete("1.0", END)
    text_box_pron.delete("1.0", END)
    text_box_exam.delete("1.0", END)
    text_box_syn.delete("1.0", END)
    text_box_ant.delete("1.0", END)
    source_entry.delete(0, END)
    source_description.set("")
    selected_source.set(None)
    text_box_turk.focus_set()

def save_to_excel(preview_data):
    """Save the data to Excel file"""
    global edit_mode, edit_row_number
    
    # Check if file exists
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        # Create new file with headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Turkish-Persian Dictionary"
        
        # Correct column order
        headers = ["ردیف", "تاریخ", "کلمه یا جمله ترکی", "تلفظ", "مثال", "مترادف", "متضاد", "منبع", "معنی فارسی"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # New file, so no edit mode
        edit_mode = False
        edit_row_number = None
    
    if edit_mode and edit_row_number is not None:
        # UPDATE existing row
        target_row = edit_row_number + 1  # Convert back to Excel row
        
        # Update the row
        ws.cell(row=target_row, column=1, value=edit_row_number)  # ردیف stays the same
        ws.cell(row=target_row, column=2, value=preview_data["date"])  # تاریخ
        ws.cell(row=target_row, column=3, value=preview_data["turkish"])  # کلمه ترکی
        ws.cell(row=target_row, column=4, value=preview_data["pronunciation"])  # تلفظ
        ws.cell(row=target_row, column=5, value=preview_data["example"])  # مثال
        ws.cell(row=target_row, column=6, value=preview_data["synonym"])  # مترادف
        ws.cell(row=target_row, column=7, value=preview_data["antonym"])  # متضاد
        ws.cell(row=target_row, column=8, value=preview_data["source"])  # منبع
        ws.cell(row=target_row, column=9, value=preview_data["persian"])  # معنی فارسی
        
        messagebox.showinfo("موفق", f"ویرایش با موفقیت ذخیره شد!\n\nردیف: {edit_row_number}\nکلمه: {preview_data['turkish']}")
        
        # Reset edit mode
        edit_mode = False
        edit_row_number = None
        
    else:
        # ADD new row
        # Find the last row with actual data in column 3
        last_data_row = 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=3).value is not None:
                last_data_row = row
        
        target_row = last_data_row + 1
        
        # Write data
        ws.cell(row=target_row, column=1, value=preview_data["row"])
        ws.cell(row=target_row, column=2, value=preview_data["date"])
        ws.cell(row=target_row, column=3, value=preview_data["turkish"])
        ws.cell(row=target_row, column=4, value=preview_data["pronunciation"])
        ws.cell(row=target_row, column=5, value=preview_data["example"])
        ws.cell(row=target_row, column=6, value=preview_data["synonym"])
        ws.cell(row=target_row, column=7, value=preview_data["antonym"])
        ws.cell(row=target_row, column=8, value=preview_data["source"])
        ws.cell(row=target_row, column=9, value=preview_data["persian"])
        
        messagebox.showinfo("موفق", f"اطلاعات با موفقیت ذخیره شد!\n\nردیف: {preview_data['row']}\nکلمه: {preview_data['turkish']}")
    
    # Save
    wb.save(EXCEL_FILE)
    wb.close()
    
    clear_form()

def check_excel_file():
    """Check if Turkish text already exists in Excel (case-insensitive)"""
    global edit_mode, edit_row_number
    
    if not os.path.exists(EXCEL_FILE):
        return None
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    turkish_text = text_box_turk.get("1.0", END).strip()
    
    if not turkish_text:
        messagebox.showwarning("اخطار", "لطفاً کلمه یا جمله ترکی را وارد کنید")
        return "empty"
    
    turkish_column = 3  # Column C
    
    for row in range(2, ws.max_row + 1):
        existing_turkish = ws.cell(row=row, column=turkish_column).value
        
        # If in edit mode, skip the row we are currently editing
        if edit_mode and edit_row_number == row - 1:
            continue
        
        if existing_turkish and existing_turkish.strip().lower() == turkish_text.lower():
            wb.close()
            return {
                "row": row - 1,  # Convert Excel row number to display number
                "turkish": existing_turkish,
                "pronunciation": ws.cell(row=row, column=4).value,
                "example": ws.cell(row=row, column=5).value,
                "synonym": ws.cell(row=row, column=6).value,
                "antonym": ws.cell(row=row, column=7).value,
                "source": ws.cell(row=row, column=8).value,
                "persian": ws.cell(row=row, column=9).value
            }
    
    wb.close()
    return None

def show_duplicate_warning(existing_data):
    """Show warning pop-up when duplicate is found"""
    popup = Toplevel(window)
    popup.title("کلمه تکراری")
    popup.geometry("550x400")
    popup.resizable(False, False)
    popup.transient(window)
    popup.grab_set()
    
    warning_label = Label(
        popup,
        text=f"⚠️ کلمه یا جمله \"{existing_data['turkish']}\" قبلاً در دیکشنری وجود دارد!",
        font=("B Nazanin", 12, "bold"),
        fg="red"
    )
    warning_label.pack(pady=20)
    
    info_frame = Frame(popup)
    info_frame.pack(pady=10, padx=20, fill=BOTH)
    
    Label(info_frame, text="اطلاعات موجود:", font=("B Nazanin", 11, "bold")).pack(anchor="e")
    Label(info_frame, text=f"معنی فارسی: {existing_data['persian']}", font=("B Nazanin", 10), justify="right").pack(anchor="e")
    
    if existing_data['pronunciation']:
        Label(info_frame, text=f"تلفظ: {existing_data['pronunciation']}", font=("B Nazanin", 10), justify="right").pack(anchor="e")
    
    if existing_data['example']:
        Label(info_frame, text=f"مثال: {existing_data['example']}", font=("B Nazanin", 10), justify="right").pack(anchor="e")
    
    if existing_data['synonym']:
        Label(info_frame, text=f"مترادف: {existing_data['synonym']}", font=("B Nazanin", 10), justify="right").pack(anchor="e")
    
    if existing_data['antonym']:
        Label(info_frame, text=f"متضاد: {existing_data['antonym']}", font=("B Nazanin", 10), justify="right").pack(anchor="e")
    
    if existing_data['source']:
        Label(info_frame, text=f"منبع: {existing_data['source']}", font=("B Nazanin", 10), justify="right").pack(anchor="e")
    
    Label(popup, text="نمی‌توان دو کلمه یا جمله یکسان داشت!", font=("B Nazanin", 10), fg="gray").pack(pady=10)
    
    button_frame = Frame(popup)
    button_frame.pack(pady=20)
    
    def cancel_action():
        popup.destroy()
    
    def rewrite_action():
        text_box_turk.delete("1.0", END)
        text_box_turk.focus_set()
        popup.destroy()
    
    def update_action():
        global edit_mode, edit_row_number
        edit_mode = True
        edit_row_number = existing_data['row']
        
        text_box_turk.delete("1.0", END)
        text_box_turk.insert("1.0", existing_data['turkish'])
        
        text_box_pron.delete("1.0", END)
        if existing_data['pronunciation']:
            text_box_pron.insert("1.0", existing_data['pronunciation'])
        
        text_box_exam.delete("1.0", END)
        if existing_data['example']:
            text_box_exam.insert("1.0", existing_data['example'])
        
        text_box_syn.delete("1.0", END)
        if existing_data.get('synonym'):
            text_box_syn.insert("1.0", existing_data['synonym'])
        
        text_box_ant.delete("1.0", END)
        if existing_data.get('antonym'):
            text_box_ant.insert("1.0", existing_data['antonym'])
        
        text_box_def.delete("1.0", END)
        if existing_data['persian']:
            text_box_def.insert("1.0", existing_data['persian'])
        
        if existing_data['source']:
            if ": " in existing_data['source']:
                source_parts = existing_data['source'].split(": ", 1)
                source_name = source_parts[0]
                source_desc = source_parts[1]
                
                if source_name in source_options:
                    selected_source.set(source_name)
                source_description.set(source_desc)
                source_entry.delete(0, END)
                source_entry.insert(0, source_desc)
            else:
                if existing_data['source'] in source_options:
                    selected_source.set(existing_data['source'])
                source_description.set("")
                source_entry.delete(0, END)
        
        popup.destroy()
        messagebox.showinfo("اطلاعات", "اطلاعات موجود در فرم بارگذاری شد. می‌توانید ویرایش کنید و سپس ذخیره کنید.")
    
    cancel_btn = Button(button_frame, text="انصراف", font=("B Nazanin", 10, "bold"), bg="#9E9E9E", fg="white", padx=20, pady=5, command=cancel_action)
    cancel_btn.grid(row=0, column=0, padx=10)
    
    rewrite_btn = Button(button_frame, text="نوشتن کلمه جدید", font=("B Nazanin", 10, "bold"), bg="#FF9800", fg="white", padx=20, pady=5, command=rewrite_action)
    rewrite_btn.grid(row=0, column=1, padx=10)
    
    update_btn = Button(button_frame, text="ویرایش کلمه موجود", font=("B Nazanin", 10, "bold"), bg="#2196F3", fg="white", padx=20, pady=5, command=update_action)
    update_btn.grid(row=0, column=2, padx=10)

def show_preview():
    """Show preview window before saving"""
    turkish_text = text_box_turk.get("1.0", END).strip()
    persian_text = text_box_def.get("1.0", END).strip()
    pronunciation_text = text_box_pron.get("1.0", END).strip()
    example_text = text_box_exam.get("1.0", END).strip()
    synonym_text = text_box_syn.get("1.0", END).strip()
    antonym_text = text_box_ant.get("1.0", END).strip()
    
    selected = selected_source.get()
    description = source_description.get().strip()
    
    if selected and selected != "None" and description:
        source_text = f"{selected}: {description}"
    elif selected and selected != "None":
        source_text = selected
    else:
        source_text = "بدون منبع"
    
    if not turkish_text:
        messagebox.showwarning("اخطار", "لطفاً کلمه یا جمله ترکی را وارد کنید")
        return
    
    if not persian_text:
        messagebox.showwarning("اخطار", "لطفاً معنی فارسی را وارد کنید")
        return
    
    row_number = get_next_row_number()
    current_date = get_current_date()
    
    preview_window = Toplevel(window)
    preview_window.title("پیش‌نمایش اطلاعات")
    preview_window.geometry("500x700")
    preview_window.resizable(False, False)
    preview_window.transient(window)
    preview_window.grab_set()
    
    preview_frame = Frame(preview_window)
    preview_frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
    
    header_frame = Frame(preview_frame)
    header_frame.pack(fill=X, pady=(0, 10))
    
    Label(header_frame, text=f"ردیف: {row_number}", font=("B Nazanin", 11, "bold"), fg="green").pack(side="right", padx=10)
    Label(header_frame, text=f"تاریخ: {current_date}", font=("B Nazanin", 11, "bold"), fg="green").pack(side="right", padx=10)
    
    sep1 = Label(preview_frame, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("B Nazanin", 10))
    sep1.pack()
    
    Label(preview_frame, text="کلمه یا جمله ترکی:", font=("B Nazanin", 11, "bold"), fg="darkblue").pack(anchor="e", pady=(10,0))
    Label(preview_frame, text=turkish_text, font=("B Nazanin", 11), fg="black", justify="right", wraplength=400).pack(anchor="e", pady=(0,5))
    
    if pronunciation_text:
        Label(preview_frame, text="تلفظ:", font=("B Nazanin", 11, "bold"), fg="darkblue").pack(anchor="e", pady=(5,0))
        Label(preview_frame, text=pronunciation_text, font=("B Nazanin", 11), fg="black", justify="right", wraplength=400).pack(anchor="e", pady=(0,5))
    
    if example_text:
        Label(preview_frame, text="مثال:", font=("B Nazanin", 11, "bold"), fg="darkblue").pack(anchor="e", pady=(5,0))
        Label(preview_frame, text=example_text, font=("B Nazanin", 11), fg="black", justify="right", wraplength=400).pack(anchor="e", pady=(0,5))
    
    if synonym_text:
        Label(preview_frame, text="مترادف:", font=("B Nazanin", 11, "bold"), fg="darkblue").pack(anchor="e", pady=(5,0))
        Label(preview_frame, text=synonym_text, font=("B Nazanin", 11), fg="black", justify="right", wraplength=400).pack(anchor="e", pady=(0,5))
    
    if antonym_text:
        Label(preview_frame, text="متضاد:", font=("B Nazanin", 11, "bold"), fg="darkblue").pack(anchor="e", pady=(5,0))
        Label(preview_frame, text=antonym_text, font=("B Nazanin", 11), fg="black", justify="right", wraplength=400).pack(anchor="e", pady=(0,5))
    
    Label(preview_frame, text="منبع:", font=("B Nazanin", 11, "bold"), fg="darkblue").pack(anchor="e", pady=(5,0))
    Label(preview_frame, text=source_text, font=("B Nazanin", 11), fg="black", justify="right", wraplength=400).pack(anchor="e", pady=(0,5))
    
    Label(preview_frame, text="معنی فارسی:", font=("B Nazanin", 11, "bold"), fg="darkblue").pack(anchor="e", pady=(5,0))
    Label(preview_frame, text=persian_text, font=("B Nazanin", 11), fg="black", justify="right", wraplength=400).pack(anchor="e", pady=(0,10))
    
    sep2 = Label(preview_frame, text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", font=("B Nazanin", 10))
    sep2.pack()
    
    button_frame = Frame(preview_window)
    button_frame.pack(pady=15)
    
    preview_data = {
        "row": row_number,
        "date": current_date,
        "turkish": turkish_text,
        "persian": persian_text,
        "pronunciation": pronunciation_text,
        "example": example_text,
        "synonym": synonym_text,
        "antonym": antonym_text,
        "source": source_text
    }
    
    def confirm_and_save():
        preview_window.destroy()
        save_to_excel(preview_data)
    
    def cancel_preview():
        preview_window.destroy()
    
    confirm_btn = Button(button_frame, text="تأیید و ذخیره", font=("B Nazanin", 10, "bold"), bg="#4CAF50", fg="white", padx=20, pady=5, command=confirm_and_save)
    confirm_btn.grid(row=0, column=0, padx=10)
    
    edit_btn = Button(button_frame, text="بازگشت و ویرایش", font=("B Nazanin", 10, "bold"), bg="#FF9800", fg="white", padx=20, pady=5, command=cancel_preview)
    edit_btn.grid(row=0, column=1, padx=10)

def on_submit():
    """Main function called when send button is clicked"""
    turkish_text = text_box_turk.get("1.0", END).strip()
    if not turkish_text:
        messagebox.showwarning("اخطار", "لطفاً کلمه یا جمله ترکی را وارد کنید")
        return
    
    duplicate = check_excel_file()
    
    if duplicate == "empty":
        return
    
    if duplicate is not None:
        show_duplicate_warning(duplicate)
    else:
        show_preview()

# ============= WIDGETS =============

# Turkish word or phrase
label_turk = Label(window, text="کلمه یا جمله ترکی", font=("B Nazanin", 12, 'bold'))
label_turk.grid(row=current_row, column=2, sticky="n")

text_box_turk = Text(window, height=1, width=50, font=("B Nazanin", 11), wrap=WORD)
text_box_turk.grid(row=current_row, column=0, columnspan=2, padx=10)
current_row += 1

# Button frame for Turkish letters
button_frame_letters = Frame(window)
button_frame_letters.grid(row=current_row, column=0, columnspan=3, pady=10)
current_row += 1

# Create buttons (2 rows x 6 columns)
for i, letter in enumerate(turkish_letters):
    row = i // 6
    col = i % 6
    btn = Button(
        button_frame_letters,
        text=letter,
        font=("Arial", 10, "bold"),
        width=4,
        height=1,
        command=lambda l=letter: insert_letter(l)
    )
    btn.grid(row=row, column=col, padx=5, pady=5)

# Info label
info_label = Label(
    window,
    text="برای درج حروف ترکی، روی دکمه مورد نظر کلیک کنید",
    font=("B Nazanin", 9),
    fg="gray"
)
info_label.grid(row=current_row, column=0, columnspan=3, pady=0)
current_row += 1

# Turkish pronunciation
label_pron = Label(window, text="تلفظ", font=("B Nazanin", 12, 'bold'))
label_pron.grid(row=current_row, column=2, sticky="n")

text_box_pron = Text(window, height=1, width=50, font=("B Nazanin", 11), wrap=WORD)
text_box_pron.grid(row=current_row, column=0, columnspan=2, padx=10)
current_row += 1

# Turkish Example
label_exam = Label(window, text="مثال", font=("B Nazanin", 12, 'bold'))
label_exam.grid(row=current_row, column=2, sticky="n")

text_box_exam = Text(window, height=1, width=50, font=("B Nazanin", 11), wrap=WORD)
text_box_exam.grid(row=current_row, column=0, columnspan=2, padx=10)
current_row += 1

# Synonym (مترادف)
label_syn = Label(window, text="مترادف", font=("B Nazanin", 12, 'bold'))
label_syn.grid(row=current_row, column=2, sticky="n")

text_box_syn = Text(window, height=1, width=50, font=("B Nazanin", 11), wrap=WORD)
text_box_syn.grid(row=current_row, column=0, columnspan=2, padx=10)
current_row += 1

# Antonym (متضاد)
label_ant = Label(window, text="متضاد", font=("B Nazanin", 12, 'bold'))
label_ant.grid(row=current_row, column=2, sticky="n")

text_box_ant = Text(window, height=1, width=50, font=("B Nazanin", 11), wrap=WORD)
text_box_ant.grid(row=current_row, column=0, columnspan=2, padx=10)
current_row += 1

# Definition box (معنی فارسی)
label_def = Label(window, text="معنی فارسی", font=("B Nazanin", 12, 'bold'))
label_def.grid(row=current_row, column=2, sticky="n")

text_box_def = Text(window, height=4, width=50, font=("B Nazanin", 11), wrap=WORD)
text_box_def.grid(row=current_row, column=0, columnspan=2, padx=10, pady=10)
current_row += 1

# ============= SOURCE SECTION - RIGHT ALIGNED =============
# Create a frame for the entire source section (right-aligned)
source_section_frame = Frame(window)
source_section_frame.grid(row=current_row, column=0, columnspan=3, pady=10, sticky="e")
current_row += 1

# "Source" label on the right side
source_label = Label(
    source_section_frame, 
    text="منبع", 
    font=("B Nazanin", 12, 'bold')
)
source_label.grid(row=0, column=3, padx=(10, 20), sticky="e")

# Frame for radio buttons
radio_frame = Frame(source_section_frame)
radio_frame.grid(row=0, column=1, padx=(0, 20))

# Create radio buttons - NONE selected
radio_buttons = []
for i, option in enumerate(source_options):
    radio_btn = Radiobutton(
        radio_frame,
        text=option,
        font=("B Nazanin", 11),
        variable=selected_source,
        value=option,
        indicatoron=1,
        selectcolor="white",
        relief="flat"
    )
    radio_btn.grid(row=0, column=i, padx=5)
    radio_buttons.append(radio_btn)

# Add New Source button
add_source_btn = Button(
    window,
    text="افزودن منبع جدید",
    font=("B Nazanin", 10, "bold"),
    bg="#2196F3",
    fg="white",
    activebackground="#1976D2",
    padx=20,
    pady=8,
    bd=1,
    relief="raised",
    cursor="hand2",
    command=add_new_source
)
add_source_btn.grid(row=current_row, column=1, pady=10)

# Remove Source button
remove_source_btn = Button(
    window,
    text="حذف منبع انتخاب شده",
    font=("B Nazanin", 10, "bold"),
    bg="#f44336",
    fg="white",
    activebackground="#d32f2f",
    padx=20,
    pady=8,
    bd=1,
    relief="raised",
    cursor="hand2",
    command=remove_selected_source
)
remove_source_btn.grid(row=current_row, column=0, padx=10)

current_row += 1

# Description section
source_desc_label = Label(
    window,
    text="توضیحات منبع:",
    font=("B Nazanin", 12, 'bold')
)
source_desc_label.grid(row=current_row, column=2, sticky="e", padx=(0, 10))

source_entry = Entry(
    window,
    textvariable=source_description,
    font=("B Nazanin", 11),
    width=50,
    relief="sunken",
    bd=1
)
source_entry.grid(row=current_row, column=0, columnspan=2, padx=10, pady=5, sticky="w")
current_row += 1

# Display the combined source
combined_source_label = Label(
    window,
    text="",
    font=("B Nazanin", 9),
    fg="blue"
)
combined_source_label.grid(row=current_row, column=0, columnspan=3, pady=5)

# Real-time preview of the combined source format
def update_preview(*args):
    selected = selected_source.get()
    description = source_description.get()
    if selected and selected != "None" and description:
        combined = f"{selected}: {description}"
        combined_source_label.config(text=f"{combined}")
    elif selected and selected != "None":
        combined_source_label.config(text=f"{selected}: ")
    else:
        combined_source_label.config(text="")

# Bind the preview update to changes
selected_source.trace_add('write', update_preview)
source_description.trace_add('write', update_preview)

current_row += 1

# Send button
send_button = Button(
    window, 
    text="ارسال کردن", 
    font=("B Nazanin", 14, "bold"),
    bg="#4CAF50",
    fg="white",
    padx=20,
    pady=0,
    command=on_submit
)
send_button.grid(row=current_row, column=1, columnspan=2, pady=10)

# Clear form button
clear_button = Button(
    window, 
    text="پاک کردن", 
    font=("B Nazanin", 14, "bold"),
    bg="#4CAF50",
    fg="white",
    padx=20,
    pady=0,
    command=clear_all
)
clear_button.grid(row=current_row, column=0, columnspan=1, pady=10)

window.geometry('550x750')
window.mainloop()